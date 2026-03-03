import io
import re
from typing import Optional, List, Dict
from datetime import time

from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.schemas import Weekday
from app.models.lesson import Lesson
from app.models.group import Group
from app.models.room import Room
from app.models.user import User


# Фиксированные временные слоты
TIME_SLOTS = [
    (time(8, 0), time(9, 35)),
    (time(9, 50), time(11, 25)),
    (time(11, 40), time(13, 15)),
    (time(13, 45), time(15, 20)),
    (time(15, 35), time(17, 10)),
    (time(17, 25), time(19, 0)),
    (time(19, 15), time(20, 50)),
]

# Регулярки
GROUP_PATTERN = re.compile(r'\b\d{2}[А-ЯA-Z]{3,}\d*\b')
TYPE_PATTERN = re.compile(r'\b(лб|лаб|лек|пр)\s*[\.\·]?\b', re.IGNORECASE)
ROOM_PATTERN = re.compile(r'[А-Яа-яA-Za-z0-9\-]{2,}$')

# Соответствие русских названий дней Weekday
WEEKDAY_NAME_MAP = {
    "понедельник": Weekday.monday,
    "вторник": Weekday.tuesday,
    "среда": Weekday.wednesday,
    "четверг": Weekday.thursday,
    "пятница": Weekday.friday,
    "суббота": Weekday.saturday,
}


# -------------------------------
#  ФУНКЦИЯ ИСПРАВЛЕНИЯ КОДИРОВКИ
# -------------------------------
def fix_encoding(s: str) -> str:
    """
    Исправляет строки, которые были прочитаны как UTF‑8,
    но на самом деле были CP1251 → UTF‑8 (кракозябры).
    """
    try:
        return s.encode("latin1").decode("cp1251")
    except:
        return s


# -------------------------------
#  ПАРСИНГ ОДНОЙ ЯЧЕЙКИ
# -------------------------------
def parse_lesson_line(line: str) -> Optional[dict]:
    line = fix_encoding(line.strip())

    if not line:
        return None

    group_match = GROUP_PATTERN.search(line)
    if not group_match:
        return None
    group_id = group_match.group()

    type_match = TYPE_PATTERN.search(line)
    if not type_match:
        return None
    lesson_type = type_match.group()

    type_index = line.find(lesson_type)
    subject = line[group_match.end():type_index].strip()

    after_type = line[type_index + len(lesson_type):].strip()
    if not after_type:
        return None

    room_match = ROOM_PATTERN.search(after_type)
    if not room_match:
        return None
    room_id = room_match.group()

    teachers_part = after_type[:room_match.start()].strip().rstrip(',')
    teachers = [t.strip() for t in teachers_part.split(',') if t.strip()]
    if not teachers:
        return None

    return {
        "group_id": group_id,
        "subject": subject,
        "type": lesson_type,
        "teachers": teachers,
        "teacher_name": teachers[0],
        "room_id": room_id,
        "subject_raw": line,
    }


# -------------------------------
#  GET/CREATE функции
# -------------------------------
def get_or_create_teacher(db: Session, full_name: str):
    full_name = fix_encoding(full_name)
    teacher = db.query(User).filter(User.full_name == full_name).first()
    if not teacher:
        teacher = User(full_name=full_name, role="teacher")
        db.add(teacher)
        db.commit()
        db.refresh(teacher)
    return teacher


def get_or_create_group(db: Session, group_code: str):
    group_code = fix_encoding(group_code)
    group = db.query(Group).filter(Group.code == group_code).first()
    if not group:
        group = Group(code=group_code, name=group_code)
        db.add(group)
        db.commit()
        db.refresh(group)
    return group


def get_or_create_room(db: Session, room_code: str):
    room_code = fix_encoding(room_code)
    room = db.query(Room).filter(Room.name == room_code).first()
    if not room:
        room = Room(name=room_code)
        db.add(room)
        db.commit()
        db.refresh(room)
    return room


# -------------------------------
#  ОПРЕДЕЛЕНИЕ КОЛОНОК ДНЕЙ НЕДЕЛИ
# -------------------------------
def detect_weekday_columns(header_row) -> List[Dict]:
    blocks = []
    current = None

    for idx, cell in enumerate(header_row):
        name = fix_encoding(str(cell)).strip().lower() if cell else ""
        if name in WEEKDAY_NAME_MAP:
            if current is not None:
                current["end"] = idx - 1
                blocks.append(current)
            current = {
                "weekday": WEEKDAY_NAME_MAP[name],
                "start": idx,
                "end": None,
            }

    if current is not None:
        current["end"] = len(header_row) - 1
        blocks.append(current)

    return blocks


# -------------------------------
#  ОСНОВНОЙ ИМПОРТ
# -------------------------------
def import_schedule_from_xlsx(db: Session, content: bytes):
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    ws = wb.active

    weekday_blocks = []
    total_lessons = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):

        # строка заголовка
        if row_idx == 2:
            weekday_blocks = detect_weekday_columns(row)
            continue

        # строка времени
        if row_idx == 3:
            continue

        first_cell = row[0]
        first_str = fix_encoding(str(first_cell)).strip() if first_cell else ""

        # строка преподавателя
        if first_str.isdigit():
            teacher_name = fix_encoding(str(row[1]).strip()) if row[1] else None
            if not teacher_name:
                continue

            teacher = get_or_create_teacher(db, teacher_name)

            for block in weekday_blocks:
                weekday = block["weekday"]
                start_col = block["start"]
                end_col = block["end"]

                for col_idx in range(start_col, end_col + 1):
                    if col_idx >= len(row):
                        continue

                    cell = row[col_idx]
                    if not cell:
                        continue

                    cell_str = fix_encoding(str(cell).replace("\n", " ").strip())
                    if not cell_str:
                        continue

                    slot_index = col_idx - start_col
                    if slot_index < 0 or slot_index >= len(TIME_SLOTS):
                        continue

                    lesson_data = parse_lesson_line(cell_str)
                    if not lesson_data:
                        continue

                    group = get_or_create_group(db, lesson_data["group_id"])
                    room = get_or_create_room(db, lesson_data["room_id"])
                    start_time, end_time = TIME_SLOTS[slot_index]

                    lesson = Lesson(
                        group_id=group.id,
                        teacher_id=teacher.id,
                        room_id=room.id,
                        subject_raw=lesson_data["subject_raw"],
                        weekday=weekday,
                        start_time=start_time,
                        end_time=end_time,
                    )
                    db.add(lesson)
                    total_lessons += 1

    db.commit()
    print(f"Импорт завершён. Создано занятий: {total_lessons}")
